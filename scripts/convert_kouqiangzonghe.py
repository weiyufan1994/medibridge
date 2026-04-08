#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
口腔综合科医生信息转换脚本
将临时txt文件转换为Excel和JSON格式
"""

import re
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

INPUT_FILE = "/home/ubuntu/medibridge/data/hospitals/上海市第九人民医院_口腔综合科_临时.txt"
OUTPUT_EXCEL = "/home/ubuntu/medibridge/data/hospitals/上海市第九人民医院_口腔综合科_医生信息_20260302.xlsx"
OUTPUT_JSON = "/home/ubuntu/medibridge/data/hospitals/上海市第九人民医院_口腔综合科_医生信息_20260302.json"

DEPT_NAME = "口腔综合科"
HOSPITAL_NAME = "上海交通大学医学院附属第九人民医院"

def parse_doctors(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    blocks = re.split(r'===医生\d+===', content)
    doctors = []
    
    for block in blocks:
        block = block.strip()
        if not block or block.startswith('#'):
            continue
        
        def get_field(name, text):
            pattern = rf'^{name}[：:]\s*(.+?)$'
            m = re.search(pattern, text, re.MULTILINE)
            return m.group(1).strip() if m else '暂无'
        
        name = get_field('姓名', block)
        if not name or name == '暂无':
            continue
        
        doctor_id = get_field('doctor_id', block)
        title = get_field('职称', block)
        hospital = get_field('医院', block)
        dept = get_field('科室', block)
        specialty_dir = get_field('专业方向', block)
        specialty = get_field('专业擅长', block)
        bio = get_field('个人简介', block)
        social = get_field('社会任职', block)
        awards = get_field('获奖荣誉', block)
        research = get_field('科研成果', block)
        rating = get_field('病友推荐度', block)
        total_patients = get_field('总患者', block)
        
        url = f"https://www.haodf.com/doctor/{doctor_id}/xinxi-jieshao.html"
        
        doctors.append({
            '姓名': name,
            '职称': title,
            '医院': hospital,
            '科室': dept,
            '专业方向': specialty_dir,
            '专业擅长': specialty,
            '个人简介': bio,
            '社会任职': social,
            '获奖荣誉': awards,
            '科研成果': research,
            '病友推荐度': rating,
            '总患者': total_patients,
            'doctor_id': doctor_id,
            '简介页URL': url,
        })
    
    return doctors

def save_excel(doctors, filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = DEPT_NAME
    
    headers = ['姓名', '职称', '医院', '科室', '专业方向', '专业擅长',
               '个人简介', '社会任职', '获奖荣誉', '科研成果',
               '病友推荐度', '总患者', 'doctor_id', '简介页URL']
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    alt_fill = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
    
    for row_idx, doctor in enumerate(doctors, 2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        for col_idx, header in enumerate(headers, 1):
            value = doctor.get(header, '暂无')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.border = thin_border
            if col_idx in [1, 2, 5, 11, 12]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
    
    col_widths = [10, 12, 28, 12, 10, 35, 50, 40, 30, 40, 12, 10, 15, 55]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    ws.row_dimensions[1].height = 25
    
    wb.save(filepath)
    print(f"Excel已保存: {filepath}")

def save_json(doctors, filepath):
    data = {
        'hospital': HOSPITAL_NAME,
        'department': DEPT_NAME,
        'scrape_date': '2026-03-03',
        'total_doctors': len(doctors),
        'doctors': doctors
    }
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"JSON已保存: {filepath}")

if __name__ == '__main__':
    doctors = parse_doctors(INPUT_FILE)
    print(f"解析到 {len(doctors)} 位医生")
    save_excel(doctors, OUTPUT_EXCEL)
    save_json(doctors, OUTPUT_JSON)
    print("=== 医生列表 ===")
    for d in doctors:
        print(f"  {d['姓名']} - {d['职称']}")
