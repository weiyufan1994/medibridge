#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
耳鼻咽喉头颈外科医生数据转换脚本
将临时txt文件转换为Excel和JSON格式
"""

import json
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# 文件路径
TXT_FILE = "/home/ubuntu/medibridge/data/hospitals/上海市第九人民医院_耳鼻咽喉头颈外科_临时.txt"
EXCEL_FILE = "/home/ubuntu/medibridge/data/hospitals/上海市第九人民医院_耳鼻咽喉头颈外科.xlsx"
JSON_FILE = "/home/ubuntu/medibridge/data/hospitals/上海市第九人民医院_耳鼻咽喉头颈外科.json"

def parse_doctors(txt_file):
    """解析临时txt文件，提取医生信息"""
    with open(txt_file, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # 按医生分割
    doctor_blocks = re.split(r'===医生\d+===', content)
    doctors = []
    
    for block in doctor_blocks:
        block = block.strip()
        if not block or block.startswith('#'):
            continue
        
        doctor = {}
        lines = block.strip().split('\n')
        
        for line in lines:
            line = line.strip()
            if line.startswith('姓名：'):
                doctor['姓名'] = line[3:].strip()
            elif line.startswith('职称：'):
                doctor['职称'] = line[3:].strip()
            elif line.startswith('doctor_id：'):
                doctor['doctor_id'] = line[10:].strip()
            elif line.startswith('医院：'):
                doctor['医院'] = line[3:].strip()
            elif line.startswith('科室：'):
                doctor['科室'] = line[3:].strip()
            elif line.startswith('专业方向：'):
                doctor['专业方向'] = line[5:].strip()
            elif line.startswith('专业擅长：'):
                doctor['专业擅长'] = line[5:].strip()
            elif line.startswith('个人简介：'):
                doctor['个人简介'] = line[5:].strip()
            elif line.startswith('社会任职：'):
                doctor['社会任职'] = line[5:].strip()
            elif line.startswith('获奖荣誉：'):
                doctor['获奖荣誉'] = line[5:].strip()
            elif line.startswith('科研成果：'):
                doctor['科研成果'] = line[5:].strip()
            elif line.startswith('病友推荐度：'):
                doctor['病友推荐度'] = line[6:].strip()
            elif line.startswith('总患者：'):
                doctor['总患者'] = line[4:].strip()
        
        if doctor.get('姓名'):
            # 生成URL
            if doctor.get('doctor_id'):
                doctor['简介URL'] = f"https://www.haodf.com/doctor/{doctor['doctor_id']}/xinxi-jieshao.html"
            # 填充缺失字段
            for field in ['职称', '专业方向', '专业擅长', '个人简介', '社会任职', '获奖荣誉', '科研成果', '病友推荐度', '总患者']:
                if field not in doctor:
                    doctor[field] = '暂无'
            doctors.append(doctor)
    
    return doctors

def save_to_excel(doctors, excel_file):
    """保存医生数据到Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "耳鼻咽喉头颈外科"
    
    # 表头
    headers = ['序号', '姓名', '职称', '医院', '科室', '专业方向', '专业擅长', 
               '个人简介', '社会任职', '获奖荣誉', '科研成果', '病友推荐度', '总患者', '简介URL']
    
    # 设置表头样式
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 写入数据
    for row, doctor in enumerate(doctors, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=doctor.get('姓名', ''))
        ws.cell(row=row, column=3, value=doctor.get('职称', ''))
        ws.cell(row=row, column=4, value=doctor.get('医院', ''))
        ws.cell(row=row, column=5, value=doctor.get('科室', ''))
        ws.cell(row=row, column=6, value=doctor.get('专业方向', ''))
        ws.cell(row=row, column=7, value=doctor.get('专业擅长', ''))
        ws.cell(row=row, column=8, value=doctor.get('个人简介', ''))
        ws.cell(row=row, column=9, value=doctor.get('社会任职', ''))
        ws.cell(row=row, column=10, value=doctor.get('获奖荣誉', ''))
        ws.cell(row=row, column=11, value=doctor.get('科研成果', ''))
        ws.cell(row=row, column=12, value=doctor.get('病友推荐度', ''))
        ws.cell(row=row, column=13, value=doctor.get('总患者', ''))
        ws.cell(row=row, column=14, value=doctor.get('简介URL', ''))
        
        # 交替行颜色
        if row % 2 == 0:
            row_fill = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
            for col in range(1, 15):
                ws.cell(row=row, column=col).fill = row_fill
        
        # 自动换行
        for col in range(1, 15):
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
    
    # 设置列宽
    col_widths = [6, 10, 15, 25, 15, 10, 40, 50, 40, 30, 40, 12, 10, 50]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    wb.save(excel_file)
    print(f"Excel文件已保存: {excel_file}")

def save_to_json(doctors, json_file):
    """保存医生数据到JSON"""
    data = {
        "医院": "上海交通大学医学院附属第九人民医院",
        "科室": "耳鼻咽喉头颈外科",
        "抓取时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "医生总数": len(doctors),
        "医生列表": doctors
    }
    
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print(f"JSON文件已保存: {json_file}")

if __name__ == "__main__":
    print("开始解析耳鼻咽喉头颈外科医生数据...")
    doctors = parse_doctors(TXT_FILE)
    print(f"共解析到 {len(doctors)} 位医生")
    
    save_to_excel(doctors, EXCEL_FILE)
    save_to_json(doctors, JSON_FILE)
    
    print("转换完成！")
    for i, d in enumerate(doctors, 1):
        print(f"{i}. {d.get('姓名', '未知')} - {d.get('职称', '未知')}")
